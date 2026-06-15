from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET


class DocumentParser:
    def extract_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if not path.exists():
            return ""
        if suffix in {".txt", ".md", ".csv", ".json"}:
            return self._read_text(path)
        if suffix == ".docx":
            return self._read_docx(path)
        if suffix == ".pdf":
            return self._read_pdf(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx(path)
        if suffix in {".jpg", ".jpeg", ".png"}:
            return f"[图片文件] {path.name}"
        return f"[暂不支持预览解析] {path.name}"

    def _read_text(self, path: Path) -> str:
        for encoding in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        return path.read_bytes().decode("utf-8", errors="ignore")

    def _read_docx(self, path: Path) -> str:
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        with ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ET.fromstring(xml)
        lines: list[str] = []
        for paragraph in root.findall(".//w:p", ns):
            text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
            if text:
                lines.append(text)
        return "\n".join(lines)

    def _read_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader
        except ImportError:
            return f"[PDF 文件，尚未安装 pypdf 解析文本] {path.name}"

        reader = PdfReader(str(path))
        pages = []
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            pages.append(f"--- Page {index} ---\n{text.strip()}")
        return "\n\n".join(pages).strip() or f"[PDF 文件，无可抽取文本] {path.name}"

    def _read_xlsx(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return f"[Excel 文件，尚未安装 openpyxl 解析文本] {path.name}"

        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            chunks.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(value.strip() for value in values):
                    chunks.append(" | ".join(values))
        workbook.close()
        return "\n".join(chunks)
